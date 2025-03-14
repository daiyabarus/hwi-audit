import csv
import datetime
import json
import os
import re
import shutil
import traceback
import xml.etree.ElementTree as ET
from enum import Enum
from typing import Iterable

from openpyxl import load_workbook


def gval(rowdata: any) -> str:
    return str(rowdata).strip().replace("None", "")


class ExclusiveEnum(str, Enum):
    @classmethod
    def list(cls):
        return list(map(lambda c: c.value, cls))

    def __str__(self) -> str:
        return str.__str__(self)


class ToGet:
    """Get Utilities for Common Use"""

    def __init__(self) -> None:
        pass

    @staticmethod
    def get_current_week() -> str:
        """Get the current week number"""
        now = datetime.datetime.now()
        return float(now.isocalendar()[1])

    @staticmethod
    def get_current_datetime() -> str:
        """Get current Date and Time YYYYMMDD_HHMMSS"""
        now = datetime.datetime.now()
        return now.strftime("%Y%m%d_%H%M%S")

    @staticmethod
    def get_current_date() -> str:
        """Get Current Date YYYYMMDD"""
        now = datetime.datetime.now()
        return now.strftime("%Y-%m-%d")

    @staticmethod
    def isfile_exists(filename: str):
        """Check is file/folder exists"""
        return os.path.exists(filename)

    @staticmethod
    def get_directory_name(path: str):
        """Get directory name from a file/dir"""
        return os.path.dirname(path)

    @staticmethod
    def get_filename(path: str):
        """Get filename from a path"""
        return os.path.basename(path)

    @staticmethod
    def combine_folder_file(folderpath: str, filename: str):
        """Combine folder and file"""
        return os.path.join(folderpath, filename)

    @staticmethod
    def open_folder(path: str):
        """Opening a folder"""
        os.startfile(path)

    @staticmethod
    def get_listdir(folderpath: str):
        """Get list of a Directory"""
        return os.listdir(folderpath)

    @staticmethod
    def get_listdir_folder(folderpath: str):
        """Get list of a folder in a path"""
        items = os.listdir(folderpath)

        folders = []
        for item in items:
            item_abspath = os.path.join(folderpath, item)
            if os.path.isdir(item_abspath):
                folders.append(item_abspath)

        return folders

    @staticmethod
    def get_listdir_with_prefix(folderpath: str, prefix: str):
        """Get list of a directory with a prefix"""
        items = os.listdir(folderpath)

        folders = []
        for item in items:
            if item.startswith(prefix):
                item_abspath = os.path.join(folderpath, item)
                if os.path.isdir(item_abspath):
                    folders.append(item_abspath)

        return folders

    @staticmethod
    def get_listfile_with_prefix(folderpath: str, prefix: str):
        """Get list of files in a directory with a specified prefix."""
        items = os.listdir(folderpath)

        files = []
        for item in items:
            if item.startswith(prefix):
                item_abspath = os.path.join(folderpath, item)
                if os.path.isfile(item_abspath):
                    files.append(item)
        return files

    @staticmethod
    def delete_file(filepath: str):
        """Delete a file"""
        try:
            if os.path.exists(filepath):
                os.remove(filepath)

            return True
        except Exception:
            return False

    @staticmethod
    def delete_ws(wb_path: str, ws_name: str):
        """Delete a worksheet in a workbook"""
        try:
            wb = load_workbook(wb_path)
            wb.remove(wb[ws_name])
            return True
        except Exception:
            return False

    @staticmethod
    def get_list_of_ws_from_wb(xlsx_path: str):
        """Get a list of worksheet from a Workbook"""
        wb = load_workbook(xlsx_path)
        sheets = []
        for sheet in wb:
            sheets.append(sheet.title)

        return sheets

    @staticmethod
    def open_sheet_by_name(xlsx_path: str, sheet_name: str):
        """Open a specific sheet by name from a Workbook."""
        wb = load_workbook(xlsx_path, data_only=True)
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        else:
            print(f"Sheet named '{sheet_name}' does not exist in the workbook.")
            return None

    @staticmethod
    def csv_to_list(csv_file: str, delimiter: str = ";") -> list:
        """Get list from a CSV file with custom delimiter"""
        list_of_mos: list = []
        with open(csv_file) as f:
            file_reader = csv.reader(f, delimiter=delimiter)
            for row in file_reader:
                if len(row) > 0:
                    list_of_mos.append(row)

        return list_of_mos

    @staticmethod
    def csv_files_to_list(csv_files: list, delimiter: str = ";") -> list:
        """Get list from CSV files with custom delimiter"""
        list_of_lists = []
        for csv_file in csv_files:
            with open(csv_file, newline="") as f:
                file_reader = csv.reader(f, delimiter=delimiter)
                for row in file_reader:
                    if len(row) > 0:
                        list_of_lists.append(row)
        return list_of_lists

    @staticmethod
    def dict_to_json(dt_info: dict, path_to_save: str):
        """Make a json file from a Dictionary"""
        try:
            with open(path_to_save, "w") as outfile:
                json.dump(dt_info, outfile)

            return True, f"json file save to: \n {path_to_save}"

        except Exception:
            errors = traceback.format_exc()
            return False, errors

    @staticmethod
    def jsonfile_to_dict(jsonfile: str) -> dict:
        """Make a Dictionary from a json file"""
        try:
            with open(jsonfile) as file:
                return dict(json.load(file))
        except Exception:
            return dict()

    @staticmethod
    def txtfile_to_list(txtpath: str) -> list:
        """Make list from a text file"""
        reader_line = []

        with open(txtpath, "r") as fl:
            liner = fl.readlines()

            for line in liner:
                cline = line.rstrip()

                reader_line.append(cline)

        return reader_line

    # INFO: Sheet to dict list
    @staticmethod
    def sheet_to_dict_list(worksheet):
        data = []
        for row in worksheet.iter_rows(values_only=True, min_row=2):
            row_data = {
                worksheet.cell(row=1, column=col).value: cell
                for col, cell in enumerate(row, start=1)
            }
            data.append(row_data)
        return data

    @staticmethod
    def copyfile(source_path: str, target_path: str) -> bool:
        """Copy a file"""
        try:
            shutil.copyfile(source_path, target_path)

            return True
        except Exception:
            return False

    @staticmethod
    def is_float(element: any) -> bool:
        """Check if an element can be converted into a float"""
        # If you expect None to be passed:
        if element is None:
            return False
        try:
            float(element)
            return True
        except ValueError:
            return False

    @staticmethod
    def get_number_fr_string(element: str) -> int:
        """Get number from any string and covert to int"""
        try:
            numbers = [s for s in element if s.isdigit()]

            print(numbers)
            numbers_str = "".join(numbers)
            return int(numbers_str)
        except TypeError:
            return 0

    @staticmethod
    def get_listdir_with_first_three_chars(folderpath: str) -> list:
        """
        Get list of directories within a given
        folderpath that start with any first three characters
        """
        items = os.listdir(folderpath)
        folders = []
        for item in items:
            if os.path.isdir(os.path.join(folderpath, item)):
                folders.append(item[:3])

        unique_folders = list(set(folders))
        return unique_folders

    @staticmethod
    def is_compare(str1: any, str2: any) -> bool:
        return str(str1).casefold() == str(str2).casefold()

    @staticmethod
    def is_in_array(str1: any, list1: Iterable):
        return str1.casefold() in (item.casefold() for item in list1)

    @staticmethod
    def is_str1_startwith_array(str1: any, list1: Iterable):
        for item in list1:
            if str(str1).startswith(str(item)):
                return True

        return False

    @staticmethod
    def starts_ends(text: str, starts: str, ends: str):
        return text.casefold().startswith(
            starts.casefold()
        ) and text.casefold().endswith(ends.casefold())

    @staticmethod
    def filter_alphanumeric_underscore(input_string: str):
        # Replace spaces with underscores
        input_string = input_string.replace(" ", "_")

        # replace double underscore to one underscore
        input_string = input_string.replace("__", "_")

        # Use a regular expression to filter out
        # non-alphanumeric characters and underscores
        filtered_string = re.sub(r"[^a-zA-Z0-9_]", "", input_string)
        return filtered_string

    @classmethod
    def from_xml(cls, xml_string):
        cell_element = ET.fromstring(xml_string)
        attributes_element = cell_element.find("attributes")

        attributes = {}
        for field_name, field_obj in cls.__dataclass_fields__[
            "attributes"
        ].type.__dataclass_fields__.items():
            element = attributes_element.find(field_obj.metadata["xml"])
            if element is not None:
                attributes[field_name] = element.text

        return cls(attributes=CellAttributes(**attributes))
